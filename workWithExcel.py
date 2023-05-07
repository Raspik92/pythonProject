import openpyxl as op
filename = "book2.xlsx"
# data_only = True читати тільки значення, без формул
wb = op.load_workbook(filename, data_only=True)
# вибираємо лист який був активний на момент відкриття
sheet = wb.active
# визнчаємо до якої строчки у нас є дані в таблиці
max_rows = sheet.max_row
print(max_rows)
for i in range(1,max_rows+1):
    print(str(i) +" " + str(sheet.cell(row=i, column=1).value)+" "+ sheet.cell(row=i, column=2).value)


print(sheet.cell(row=7,column=2).value)